import os
import requests
from openpyxl import load_workbook
from tqdm import tqdm
import mimetypes
import re
import traceback
import json

# 读取文件名，调用dify读取文件进行对比，将结果保存到excel

# 配置项
API_KEY = 'app-YZHlycEoDIa6ov4lLGoIlK9C'  
USER_ID = 'abc-123123'
EXCEL_PATH = r'D:\Download\dayu5000\所有匹配结果-2025-07-09-2-json.xlsx'

UPLOAD_URL = 'https://dev-dify.hysz.co/v1/files/upload'
CHAT_URL = 'https://dev-dify.hysz.co/v1/chat-messages'


# === 文件上传 ===
def upload_file(filepath):
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在：{filepath}")
    
    headers = {'Authorization': f'Bearer {API_KEY}'}
    data = {'user': USER_ID}
    
    # 自动识别 MIME 类型
    mime_type, _ = mimetypes.guess_type(filepath)
    if mime_type is None:
        mime_type = 'application/octet-stream'  # 默认类型，防止识别失败
    
    with open(filepath, 'rb') as f:
        files = {
            'file': (
                os.path.basename(filepath),
                f,
                mime_type
            )
        }
        response = requests.post(UPLOAD_URL, headers=headers, files=files, data=data)
    
    response.raise_for_status()
    return response.json()['id']

# === 文件对比 ===
def compare_files(file1_id, file2_id):
    headers = {'Authorization': f'Bearer {API_KEY}', 'Content-Type': 'application/json'}
    payload = {
        "inputs": {
            # "fileupload": [file1_id, file2_id],
            "fileupload": [
                    {"type": "document", "transfer_method": "local_file", "upload_file_id": file1_id},
                    {"type": "document", "transfer_method": "local_file", "upload_file_id": file2_id}
                ],
            # "file_num": ["1", "2"]
        },
        "query": "合同对比",
        "response_mode": "blocking",
        "conversation_id": "",
        "user": USER_ID,
        # "files": [
        #     {"type": "document", "transfer_method": "local_file", "upload_file_id": file1_id},
        #     {"type": "document", "transfer_method": "local_file", "upload_file_id": file2_id}
        # ]
    }
    resp = requests.post(CHAT_URL, headers=headers, json=payload)
    resp.raise_for_status()
    return resp.json().get('answer', '无返回')

# === 处理 Excel ===
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# 标题行
headers = [cell.value for cell in ws[1]]

# 获取已有列位置
col_map = {}
required_cols = ['合同名称', '合同最终附件路径', '最相似合同创建附件路径']
for name in required_cols:
    col_map[name] = headers.index(name) + 1

# 添加缺失列
def ensure_column(title):
    if title in headers:
        return headers.index(title) + 1
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=title)
    headers.append(title)
    return col

# 新增字段
col_final_id = ensure_column('最终附件file_id')
col_create_id = ensure_column('创建附件file_id')
col_result = ensure_column('对比结果')
col_status = ensure_column('状态')

# 遍历数据行
for row in tqdm(range(2, ws.max_row + 1), desc="处理合同对比"):

    contract = ws.cell(row=row, column=col_map['合同名称']).value
    final_path = ws.cell(row=row, column=col_map['合同最终附件路径']).value
    create_path = ws.cell(row=row, column=col_map['最相似合同创建附件路径']).value
    # 路径为空，跳过
    if not final_path or not create_path:
        ws.cell(row=row, column=col_status, value='路径为空，跳过')
        continue
    
    final_id = ws.cell(row=row, column=col_final_id).value
    create_id = ws.cell(row=row, column=col_create_id).value
    compare_result = ws.cell(row=row, column=col_result).value
    print("正在处理项目：" + contract + "，合同：" + final_path)
    try:
        # 上传文件（如果没上传过）
        if not final_id:
            final_id = upload_file(final_path)
            ws.cell(row=row, column=col_final_id, value=final_id)
            wb.save(EXCEL_PATH)

        if not create_id:
            create_id = upload_file(create_path)
            ws.cell(row=row, column=col_create_id, value=create_id)
            wb.save(EXCEL_PATH)

        # 如果对比结果已存在，跳过
        if compare_result:
            print("项目："+ contract + "，合同：" + final_id + "，合同已对比，跳过")
            continue

        # 调用对比接口
        result = compare_files(final_id, create_id)
       
        cleaned_result = re.sub(r'<think>.*?</think>', '', result, flags=re.DOTALL).strip()
        print("项目："+ contract + "，合同：" + final_id + "，对比的结果为：" + cleaned_result)
        ws.cell(row=row, column=col_result, value=cleaned_result)
        
        # 尝试解析 JSON 内容
        try:
            result_json = json.loads(cleaned_result) if isinstance(cleaned_result, str) else cleaned_result
            if isinstance(result_json, dict):
                # 添加 合同总金额 列
                col_amount = ensure_column('合同总金额')
                ws.cell(row=row, column=col_amount, value=result_json.get('contractTotalAmount', ''))

                # 添加 不同处数量 列
                col_diff_count = ensure_column('不同处数量')
                ws.cell(row=row, column=col_diff_count, value=result_json.get('diffCount', ''))

                # 动态添加 每个差异说明
                diffs = result_json.get('contractDiffAnalysis', [])
                for idx, item in enumerate(diffs, start=1):
                    col_diff_n = ensure_column(f'不同处明细文本_{idx}')
                    ws.cell(row=row, column=col_diff_n, value=item.get('description', ''))

        except Exception as json_err:
            ws.cell(row=row, column=col_status, value=f'JSON解析失败:\n{str(json_err)}')
        else:
            ws.cell(row=row, column=col_status, value='已完成')

    except Exception as e:
        error_msg = traceback.format_exc()
        ws.cell(row=row, column=col_status, value=f'出错:\n{error_msg}')

    # 每处理一行就保存（断点续传核心）
    wb.save(EXCEL_PATH)

print("✅ 所有合同已处理完成！")
